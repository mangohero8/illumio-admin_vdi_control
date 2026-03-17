#!/usr/bin/env python3
"""
MON.C9.9 — Admin VDI Control Test (Automated)
Illumio Admin VDI Restrictions Compliance Scanner

Automates the manual control test procedure from the Admin VDI Control
Test Guide v1.2. Scans extra-scope rulesets for rules allowing traffic
over admin VDI ports from non-permitted sources.

Modes:
  --csv <file>   CSV fallback (pre-exported rule search from Illumio UI)
  (default)      Direct API pull from Illumio PCE

Decision Filter Chain:
  1. Strip non-end-user policy objects (permitted IPLs, A-*, E-*, etc.)
     — preserves A-END_USER_COMPUTE_* (EUC) for review
  2. No sources remain → Excluded (all permitted)
  3. EUC source present → Requires Review
  4. Non-excluded IPL- present → Requires Review
  5. App/Env labels only (safety net) → Excluded
  6. Only label groups remain → Requires Review (edge case)
  7. Everything else → Non-Compliant

Version: 2.1.0
Author: Cybersecurity Tech Ops — Illuminati Team
"""

__version__ = "2.1.0"

import os, sys, csv, hashlib, getpass, logging, platform, argparse
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import yaml, requests
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

CONTROL_ID = "MON.C9.9"
CONTROL_NAME = "Admin VDI Restrictions"
DEFAULT_CONFIG_PATH = "config.yaml"

# Type prefixes present in Illumio SaaS CSV exports
SAAS_TYPE_PREFIXES = ["IP Lists: ", "Label Group: ", "Label: "]

# Scope key prefixes to strip everywhere
SCOPE_KEY_PREFIXES = ["app:", "env:", "loc:", "role:"]

# Decision categories
D_EXCLUDE_ALL_PERMITTED = "Exclude – No Remaining Sources With Access"
D_EXCLUDE_APP_TO_APP    = "Exclude – Application to Application Traffic is Permitted"
D_REVIEW_EUC            = "Keep for Review – Contains A-END_USER_COMPUTE_(EUC)"
D_REVIEW_IPL            = "Keep for Review – Contains an IP List"
D_REVIEW_EDGE           = "Keep for Review – Edge Case (Label Group Only)"
D_NON_COMPLIANT         = "Non-Compliant – Unpermitted Source on Restricted Port"


# =============================================================================
# HELPERS
# =============================================================================

def strip_scope_prefixes(text: str) -> str:
    for p in SCOPE_KEY_PREFIXES:
        text = text.replace(p, "")
    return text.strip()

def strip_all_type_prefixes(text: str) -> str:
    for p in SAAS_TYPE_PREFIXES:
        text = text.replace(p, "")
    return text.strip()

def matches_pattern(value: str, patterns: list[str]) -> bool:
    for p in patterns:
        if p.endswith("*") and value.startswith(p[:-1]):
            return True
        elif value == p:
            return True
    return False


# =============================================================================
# CONFIG
# =============================================================================

def load_config(path: str = DEFAULT_CONFIG_PATH) -> dict:
    config = {}
    if Path(path).exists():
        with open(path) as f:
            config = yaml.safe_load(f) or {}
        logging.info(f"Loaded config: {path}")
    else:
        logging.warning(f"Config not found: {path}")

    pce = config.get("pce", {})
    config["_pce"] = {
        "fqdn":     os.environ.get("ILLUMIO_FQDN",     pce.get("fqdn", "")),
        "port":     os.environ.get("ILLUMIO_PORT",      str(pce.get("port", "8443"))),
        "org_id":   os.environ.get("ILLUMIO_ORG_ID",    str(pce.get("org_id", "1"))),
        "api_user": os.environ.get("ILLUMIO_API_USER",  pce.get("api_user", "")),
        "api_key":  os.environ.get("ILLUMIO_API_KEY",   pce.get("api_key", "")),
    }
    missing = [k for k in ("fqdn", "api_user", "api_key") if not config["_pce"][k]]
    if missing:
        raise ValueError(f"Missing PCE config: {', '.join(missing)}")
    return config


# =============================================================================
# LOGGING
# =============================================================================

class LogCapture(logging.Handler):
    def __init__(self):
        super().__init__()
        self.records = []
    def emit(self, record):
        self.records.append({
            "timestamp": datetime.fromtimestamp(record.created).strftime("%Y-%m-%d %H:%M:%S"),
            "level": record.levelname,
            "message": self.format(record),
        })

def setup_logging(log_file="vdi_control_test.log"):
    cap = LogCapture()
    cap.setLevel(logging.DEBUG)
    cap.setFormatter(logging.Formatter("%(message)s"))
    fh = logging.FileHandler(log_file, mode="a")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    root = logging.getLogger()
    root.setLevel(logging.DEBUG)
    root.addHandler(fh)
    root.addHandler(ch)
    root.addHandler(cap)
    return cap


# =============================================================================
# ILLUMIO API CLIENT
# =============================================================================

class IllumioClient:
    """REST client for Illumio PCE API v2."""

    # Use a large max_results to fetch all items in a single request,
    # matching the approach used in the proven working script.
    MAX_RESULTS = 100000

    def __init__(self, fqdn, port, org_id, api_user, api_key):
        self.base = f"https://{fqdn}:{port}/api/v2/orgs/{org_id}"
        self.auth = HTTPBasicAuth(api_user, api_key)
        self.s = requests.Session()
        self.s.verify = False
        self.s.auth = self.auth
        self.s.headers["Accept"] = "application/json"
        self.fqdn = fqdn
        self.org_id = org_id

    def _fetch(self, ep, params=None):
        """GET with retry (3 attempts) and max_results=100000."""
        if params is None:
            params = {}
        params.setdefault("max_results", self.MAX_RESULTS)

        url = f"{self.base}{ep}"
        for attempt in range(1, 4):
            try:
                logging.info(f"Fetching items from URL: {url}?max_results={params['max_results']}")
                r = self.s.get(url, params=params, timeout=120)
                r.raise_for_status()
                data = r.json()

                # Handle both list and dict response formats
                items = data if isinstance(data, list) else data.get("items", data.get("results", []))
                logging.info(f"Successfully fetched {len(items)} items from {url}")
                return items
            except Exception as e:
                logging.error(f"API error {url} (attempt {attempt}): {e}")
                if attempt == 3: raise

    def get_services(self):
        logging.info("Building service map...")
        return self._fetch("/sec_policy/active/services")

    def get_rulesets(self):
        logging.info("Fetching all rulesets...")
        return self._fetch("/sec_policy/active/rule_sets")

    def get_labels(self):
        logging.info("Building label map...")
        return self._fetch("/labels")

    def get_ip_lists(self):
        logging.info("Building IP list map...")
        return self._fetch("/sec_policy/active/ip_lists")

    def get_label_groups(self):
        logging.info("Building label group map...")
        return self._fetch("/sec_policy/active/label_groups")


# =============================================================================
# SERVICE RESOLUTION
# =============================================================================

def find_restricted_services(services, restricted_ports):
    """Returns dict: service_href -> service_name for services containing restricted ports."""
    port_nums = [rp["port"] for rp in restricted_ports]
    restricted = {}
    for svc in services:
        for pi in svc.get("service_ports", []):
            p, tp = pi.get("port"), pi.get("to_port")
            if p and tp:
                if any(p <= rp <= tp for rp in port_nums):
                    restricted[svc["href"]] = svc["name"]; break
            elif p and p in port_nums:
                restricted[svc["href"]] = svc["name"]; break
    logging.info(f"Restricted services: {len(restricted)}")
    return restricted


def check_csv_services_restricted(services_str, restricted_svc_names, restricted_ports):
    """Check if a CSV services field contains restricted services/ports."""
    port_nums = [rp["port"] for rp in restricted_ports]
    for svc in services_str.split("; "):
        svc = svc.strip()
        if not svc:
            continue
        if svc in restricted_svc_names:
            return True
        # Direct port refs like "3389 TCP", "7389 UDP"
        for rp in port_nums:
            if svc.startswith(str(rp)):
                return True
    return False


# =============================================================================
# HREF LOOKUP (API mode)
# =============================================================================

def build_href_lookup(labels, ip_lists, label_groups=None):
    lookup = {}
    for l in labels:
        lookup[l.get("href", "")] = {"name": l.get("value", l.get("name", "?")), "key": l.get("key", ""), "type": "label"}
    for ip in ip_lists:
        lookup[ip.get("href", "")] = {"name": ip.get("name", "?"), "key": "ip_list", "type": "ip_list"}
    for lg in (label_groups or []):
        lookup[lg.get("href", "")] = {"name": lg.get("name", "?"), "key": lg.get("key", ""), "type": "label_group"}
    return lookup


def resolve_actors(actors, lookup):
    resolved = []
    for a in actors:
        if not isinstance(a, dict):
            resolved.append({"name": str(a), "type": "unknown"}); continue
        if "actors" in a:
            resolved.append({"name": a["actors"], "type": "actors"})
        elif "label" in a:
            info = lookup.get(a["label"].get("href", ""), {"name": "?", "type": "label"})
            resolved.append({"name": info["name"], "type": "label"})
        elif "label_group" in a:
            info = lookup.get(a["label_group"].get("href", ""), {"name": "?", "type": "label_group"})
            resolved.append({"name": info["name"], "type": "label_group"})
        elif "ip_list" in a:
            info = lookup.get(a["ip_list"].get("href", ""), {"name": "?", "type": "ip_list"})
            resolved.append({"name": info["name"], "type": "ip_list"})
        elif "workload" in a:
            resolved.append({"name": a["workload"].get("href", "?"), "type": "workload"})
        else:
            resolved.append({"name": str(a), "type": "unknown"})
    return resolved


def check_api_services_restricted(ingress_services, restricted_svc_map, restricted_ports):
    port_nums = [rp["port"] for rp in restricted_ports]
    names, is_restricted = [], False
    for svc in ingress_services:
        if not isinstance(svc, dict): names.append(str(svc)); continue
        if "port" in svc:
            p, tp, proto = svc["port"], svc.get("to_port"), svc.get("proto", "")
            ps = {6: "TCP", 17: "UDP"}.get(proto, str(proto))
            names.append(f"{ps}/{p}-{tp}" if tp else f"{ps}/{p}")
            if tp:
                if any(p <= rp <= tp for rp in port_nums): is_restricted = True
            elif p in port_nums: is_restricted = True
        elif "href" in svc:
            h = svc["href"]
            if h in restricted_svc_map: is_restricted = True; names.append(restricted_svc_map[h])
            else: names.append(h.split("/")[-1])
    return names, is_restricted


# =============================================================================
# PARSE CSV SOURCES/DESTINATIONS
# =============================================================================

def parse_csv_field(raw: str) -> list[dict]:
    """Parse a combined SaaS export field (Sources or Destinations) into structured list."""
    items = []
    for part in raw.split("; "):
        part = part.strip()
        if not part:
            continue
        if part.startswith("IP Lists: "):
            items.append({"name": part[len("IP Lists: "):], "type": "ip_list"})
        elif part.startswith("Label Group: "):
            items.append({"name": part[len("Label Group: "):], "type": "label_group"})
        elif part.startswith("Label: "):
            items.append({"name": part[len("Label: "):], "type": "label"})
        elif part == "All Workloads":
            items.append({"name": "All Workloads", "type": "actors"})
        else:
            items.append({"name": part, "type": "unknown"})
    return items


# =============================================================================
# DECISION FILTER ENGINE
# =============================================================================

class DecisionEngine:
    """
    Full decision filter chain from MON.C9.9 guide v1.2:
      1. Strip non-end-user policy objects (preserving EUC)
      2. No sources remain → Excluded
      3. EUC present → Review
      4. Non-excluded IPL- → Review
      5. All remaining A-*/E-* → Excluded (safety net / app-to-app)
      6. Only label groups → Review (edge case)
      7. Everything else → Non-Compliant
    """
    def __init__(self, config):
        c = config.get("compliance", {})
        self.excluded_ipls = c.get("excluded_sources", [])
        self.excluded_labels = c.get("excluded_labels", [])
        self.excluded_lgs = c.get("excluded_label_groups", [])
        self.euc_patterns = c.get("euc_patterns", ["A-END_USER_COMPUTE_*"])

    def run(self, sources: list[dict]) -> tuple[str, str, list[dict]]:
        remaining, removed = [], []

        for src in sources:
            name, stype = src["name"], src["type"]

            # Preserve EUC BEFORE A-* exclusion
            if any(matches_pattern(name, [p]) for p in self.euc_patterns):
                remaining.append(src); continue

            # Excluded IPLs
            if stype == "ip_list" and matches_pattern(name, self.excluded_ipls):
                removed.append(name); continue

            # Excluded labels (A-*, E-*, R-METTLE-CI, etc.)
            if stype == "label" and matches_pattern(name, self.excluded_labels):
                removed.append(name); continue

            # Excluded label groups
            if stype == "label_group" and matches_pattern(name, self.excluded_lgs):
                removed.append(name); continue

            # All Workloads actor
            if stype == "actors" and name == "All Workloads":
                removed.append(name); continue

            remaining.append(src)

        if removed:
            logging.debug(f"  Stripped: {', '.join(removed)}")

        # Step 2
        if not remaining:
            return D_EXCLUDE_ALL_PERMITTED, f"All sources permitted: {', '.join(removed)}", []

        # Step 3 — EUC
        euc = [s for s in remaining if any(matches_pattern(s["name"], [p]) for p in self.euc_patterns)]
        if euc:
            return D_REVIEW_EUC, f"EUC source(s): {', '.join(s['name'] for s in euc)} — end-user laptops, review for access outside IPL-ADMIN_VDI", remaining

        # Step 4 — Non-excluded IPL-
        ipls = [s for s in remaining if s["type"] == "ip_list" or s["name"].startswith("IPL-")]
        if ipls:
            return D_REVIEW_IPL, f"Non-excluded IP List(s): {', '.join(s['name'] for s in ipls)}", remaining

        # Step 5 — All remaining are A-*/E-* (app-to-app / env traffic)
        if all(s["name"].startswith("A-") or s["name"].startswith("E-") for s in remaining):
            return D_EXCLUDE_APP_TO_APP, f"App/Env traffic: {', '.join(s['name'] for s in remaining)}", []

        # Step 6 — Only label groups
        if all(s["type"] == "label_group" for s in remaining):
            return D_REVIEW_EDGE, f"Only label groups: {', '.join(s['name'] for s in remaining)}", remaining

        # Step 7 — Non-compliant
        details = []
        for s in remaining:
            if s["name"].startswith("R-"):
                details.append(f"{s['name']} (role label)")
            elif s["type"] == "ip_list":
                details.append(f"{s['name']} (IP list)")
            else:
                details.append(f"{s['name']} ({s['type']})")
        return D_NON_COMPLIANT, f"Non-permitted source(s): {', '.join(details)}", remaining


# =============================================================================
# API MODE
# =============================================================================

def extract_scope(ruleset, lookup):
    parts = []
    for ss in ruleset.get("scopes", []):
        if isinstance(ss, list):
            for e in ss:
                if isinstance(e, dict):
                    info = lookup.get(e.get("label", {}).get("href", ""), {"name": "?"})
                    parts.append(info["name"])
    return " | ".join(parts) or "Unscoped"

def is_epd_scope(ruleset, lookup):
    """Check if ruleset is scoped to E-PD (Production). Strict match."""
    for ss in ruleset.get("scopes", []):
        if isinstance(ss, list):
            for e in ss:
                if isinstance(e, dict):
                    info = lookup.get(e.get("label", {}).get("href", ""), {})
                    if info.get("key") == "env" and info.get("name") == "E-PD":
                        return True
    return False

def process_api(client, config):
    comp = config.get("compliance", {})
    rp = comp.get("restricted_ports", [{"port":22,"proto":6},{"port":23,"proto":6},{"port":3389,"proto":6},{"port":3389,"proto":17},{"port":7389,"proto":6},{"port":7389,"proto":17}])

    svcs = client.get_services()
    labels = client.get_labels()
    ips = client.get_ip_lists()
    lgs = client.get_label_groups()
    rsets = client.get_rulesets()

    lookup = build_href_lookup(labels, ips, lgs)
    rsvc = find_restricted_services(svcs, rp)
    engine = DecisionEngine(config)

    nc, rv, ex = [], [], []
    stats = dict(total_rulesets=len(rsets), production_rulesets=0, total_rules=0, extra_scope=0,
                 restricted_port_rules=0, non_compliant=0, needs_review=0, excluded=0,
                 skip_non_prod=0, skip_intra=0, skip_disabled=0, skip_no_restricted=0)

    for rs in rsets:
        rs_name = rs.get("name", "Unknown")
        scope = extract_scope(rs, lookup)
        rules = rs.get("rules", [])

        # Strict E-PD scope filter
        if not is_epd_scope(rs, lookup):
            stats["skip_non_prod"] += 1
            logging.info(f"Rule skipped: not in E-PD scope.")
            continue

        stats["production_rulesets"] += 1
        logging.info(f"Processing ruleset: {rs_name} | {scope} with {len(rules)} rules.")

        for rule in rules:
            stats["total_rules"] += 1

            if not rule.get("unscoped_consumers", False):
                stats["skip_intra"] += 1; continue
            stats["extra_scope"] += 1

            if not rule.get("enabled", True):
                stats["skip_disabled"] += 1; continue

            sources = resolve_actors(rule.get("consumers", []), lookup)
            dests = resolve_actors(rule.get("providers", []), lookup)
            svc_names, has_restricted = check_api_services_restricted(rule.get("ingress_services", []), rsvc, rp)

            src_str = "; ".join(s["name"] for s in sources)
            dst_str = "; ".join(d["name"] for d in dests)
            svc_str = "; ".join(svc_names)

            # Log special cases like working script does
            if any(s["name"] == "All Workloads" and s["type"] == "actors" for s in sources):
                logging.info(f"Rule source includes All Workloads (ams).")

            entry = {"ruleset": rs_name, "scopes": scope, "rule_href": rule.get("href",""),
                     "sources": src_str, "sources_remaining": "", "destinations": dst_str,
                     "services": svc_str, "decision": "", "reason": ""}

            if not has_restricted:
                stats["skip_no_restricted"] += 1
                logging.info("Rule skipped: no restricted services.")
                entry["decision"] = "N/A – No Restricted Ports"
                ex.append(entry); stats["excluded"] += 1; continue

            stats["restricted_port_rules"] += 1
            dec, reason, rem = engine.run(sources)
            entry["decision"] = dec
            entry["reason"] = reason
            entry["sources_remaining"] = "; ".join(s["name"] for s in rem) if rem else "(none)"

            if dec.startswith("Exclude"):
                if not rem:
                    logging.warning(f"Rule with blank sources or destinations in ruleset: {rs_name} | {scope}")
                else:
                    logging.info(f"Rule skipped: sources matches exclude pattern.")
                ex.append(entry); stats["excluded"] += 1
            elif dec.startswith("Keep"):
                logging.info(f"Rule flagged for review: {dec}")
                rv.append(entry); stats["needs_review"] += 1
            else:
                logging.warning(f"NON-COMPLIANT rule in {rs_name}: {reason}")
                nc.append(entry); stats["non_compliant"] += 1

    return nc, rv, ex, stats


# =============================================================================
# CSV MODE
# =============================================================================

def process_csv(csv_path, config, client=None):
    comp = config.get("compliance", {})
    rp = comp.get("restricted_ports", [{"port":22,"proto":6},{"port":23,"proto":6},{"port":3389,"proto":6},{"port":3389,"proto":17},{"port":7389,"proto":6},{"port":7389,"proto":17}])

    # Build restricted service name set from API if available
    rsvc_names = set()
    if client:
        try:
            svcs = client.get_services()
            rsvc_names = set(find_restricted_services(svcs, rp).values())
        except Exception as e:
            logging.warning(f"Could not fetch services for CSV mode: {e}")
    # Always include known restricted service names
    rsvc_names.update(["SVC-SSH", "SVC-TELNET", "SVC-MS-RDP"])

    with open(csv_path) as f:
        rows = list(csv.DictReader(f))
    logging.info(f"CSV: {len(rows)} rules from {csv_path}")

    engine = DecisionEngine(config)
    nc, rv, ex = [], [], []
    stats = dict(total_rulesets="N/A (CSV)", production_rulesets="N/A (CSV)",
                 total_rules=len(rows), extra_scope=len(rows),
                 restricted_port_rules=0, non_compliant=0, needs_review=0, excluded=0,
                 skip_non_prod=0, skip_intra=0, skip_disabled=0, skip_no_restricted=0)

    for row in rows:
        # Parse fields — handle both SaaS combined format and split column format
        raw_sources = row.get("Sources", row.get("sources", ""))
        raw_dests = row.get("Destinations", row.get("destinations", ""))
        raw_services = row.get("Destination Services", row.get("services", ""))
        raw_scopes = row.get("Scopes", row.get("scopes", row.get("ruleset_scope", "")))
        raw_ruleset = row.get("Ruleset", row.get("ruleset_name", ""))

        if not raw_ruleset:
            continue

        # Parse structured sources/destinations
        sources = parse_csv_field(raw_sources)
        dests = parse_csv_field(raw_dests)

        # Strip all prefixes for display
        scopes_clean = strip_all_type_prefixes(strip_scope_prefixes(raw_scopes))
        src_display = strip_all_type_prefixes(raw_sources)
        dst_display = strip_all_type_prefixes(raw_dests)

        # Check for restricted services/ports
        if not check_csv_services_restricted(raw_services, rsvc_names, rp):
            stats["skip_no_restricted"] += 1
            ex.append({"ruleset": raw_ruleset, "scopes": scopes_clean, "rule_href": "N/A (CSV)",
                       "sources": src_display, "sources_remaining": "",
                       "destinations": dst_display, "services": raw_services,
                       "decision": "N/A – No Restricted Ports", "reason": ""})
            stats["excluded"] += 1
            continue

        stats["restricted_port_rules"] += 1
        dec, reason, rem = engine.run(sources)
        rem_display = "; ".join(s["name"] for s in rem) if rem else "(none)"

        entry = {"ruleset": raw_ruleset, "scopes": scopes_clean, "rule_href": "N/A (CSV)",
                 "sources": src_display, "sources_remaining": rem_display,
                 "destinations": dst_display, "services": raw_services,
                 "decision": dec, "reason": reason}

        if dec.startswith("Exclude"):
            ex.append(entry); stats["excluded"] += 1
        elif dec.startswith("Keep"):
            rv.append(entry); stats["needs_review"] += 1
        else:
            nc.append(entry); stats["non_compliant"] += 1

    return nc, rv, ex, stats


# =============================================================================
# EXCEL REPORT
# =============================================================================

HF = Font(bold=True, color="FFFFFF", size=11)
F_RED = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
F_AMB = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
F_GRN = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
F_GRY = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
F_BLU = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
BG_PASS = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BG_FAIL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
BG_COND = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
HDRS = ["Ruleset", "Scopes", "Rule HREF", "Sources (Original)", "Sources (After Filter)",
        "Destinations", "Destination Services", "Decision Filter", "Decision Reason"]

def _style_hdr(ws, row, fill):
    for c in ws[row]:
        c.font = HF; c.fill = fill; c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = THIN

def _auto_w(ws, mx=55):
    for i, col in enumerate(ws.columns, 1):
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(w + 2, mx)

def _write_sheet(ws, rules, fill, empty_msg):
    ws.append(HDRS); _style_hdr(ws, 1, fill)
    if not rules:
        ws.append([empty_msg] + [""] * (len(HDRS)-1))
        ws.cell(2, 1).font = Font(italic=True, color="808080")
    else:
        for r in rules:
            ws.append([r.get("ruleset",""), r.get("scopes",""), r.get("rule_href",""),
                       r.get("sources",""), r.get("sources_remaining",""),
                       r.get("destinations",""), r.get("services",""),
                       r.get("decision",""), r.get("reason","")])
    _auto_w(ws)

def fmt_ports(rp_list):
    parts = []
    for rp in rp_list:
        if isinstance(rp, dict):
            proto = {6:"TCP",17:"UDP"}.get(rp.get("proto",6), str(rp.get("proto","?")))
            parts.append(f"{rp['port']}/{proto}")
        else: parts.append(str(rp))
    return ", ".join(parts)

def generate_report(nc, rv, ex, stats, logs, config, out_path, mode):
    wb = Workbook()
    pce = config["_pce"]
    comp = config.get("compliance", {})
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    # -- Audit Summary --
    ws = wb.active; ws.title = "Audit Summary"
    nc_ct, rv_ct = stats.get("non_compliant",0), stats.get("needs_review",0)

    if nc_ct == 0 and rv_ct == 0: result = "PASS — No non-compliant or review-pending rules"
    elif nc_ct == 0: result = f"CONDITIONAL — {rv_ct} rule(s) require manual review"
    else: result = f"FAIL — {nc_ct} non-compliant, {rv_ct} require review"

    summary = [
        (f"{CONTROL_ID} — {CONTROL_NAME} — Control Test Report", ""),
        ("",""),
        ("EXECUTION METADATA",""),
        ("Script Version", __version__),
        ("Control", f"{CONTROL_ID} — {CONTROL_NAME}"),
        ("Timestamp", now),
        ("Executed By", getpass.getuser()),
        ("Hostname", platform.node()),
        ("PCE Target", pce["fqdn"]),
        ("Org ID", pce["org_id"]),
        ("Data Source", "Illumio API (Direct)" if mode == "api" else f"CSV Fallback: {mode}"),
        ("",""),
        ("SCAN SCOPE",""),
        ("Environment", "Production (E-PD)"),
        ("Rule Scope", "Extra-Scope Only"),
        ("Rule Status", "Enabled Only"),
        ("Restricted Ports", fmt_ports(comp.get("restricted_ports", []))),
        ("",""),
        ("SCAN STATISTICS",""),
        ("Total Rulesets", stats.get("total_rulesets","N/A")),
        ("Production Rulesets", stats.get("production_rulesets","N/A")),
        ("Skipped (Non-Prod)", stats.get("skip_non_prod",0)),
        ("Total Rules Scanned", stats.get("total_rules",0)),
        ("Extra-Scope Rules", stats.get("extra_scope",0)),
        ("Skipped (Intra-Scope)", stats.get("skip_intra",0)),
        ("Skipped (Disabled)", stats.get("skip_disabled",0)),
        ("Skipped (No Restricted)", stats.get("skip_no_restricted",0)),
        ("Rules with Restricted Ports", stats.get("restricted_port_rules",0)),
        ("",""),
        ("DECISION RESULTS",""),
        ("Non-Compliant", nc_ct),
        ("Requires Manual Review", rv_ct),
        ("Excluded / Compliant", stats.get("excluded",0)),
        ("",""),
        ("CONTROL TEST RESULT", result),
        ("",""),
        ("DECISION FILTER CHAIN",""),
        ("Step 1", "Strip permitted sources: excluded IPLs, A-*, E-*, All Workloads, specific labels/LGs (preserve EUC)"),
        ("Step 2", "No sources remain → Excluded (all sources were permitted policy objects)"),
        ("Step 3", "A-END_USER_COMPUTE_(EUC) present → Requires Review (end-user laptops)"),
        ("Step 4", "Non-excluded IPL- sources remain → Requires Review (non-admin IP list)"),
        ("Step 5", "All remaining A-*/E-* → Excluded (app-to-app / environment traffic)"),
        ("Step 6", "Only label groups remain → Requires Review (edge case, may apply broadly)"),
        ("Step 7", "Everything else (R-* role labels, other) → Non-Compliant"),
        ("",""),
        ("EXCLUDED SOURCES (config.yaml)",""),
    ]
    for s in comp.get("excluded_sources",[]): summary.append((f"  IPL: {s}",""))
    for s in comp.get("excluded_labels",[]): summary.append((f"  Label: {s}",""))
    for s in comp.get("excluded_label_groups",[]): summary.append((f"  Label Group: {s}",""))
    for s in comp.get("euc_patterns",[]): summary.append((f"  EUC (Review): {s}",""))

    for r in summary: ws.append(r)

    ws["A1"].font = Font(bold=True, size=14)
    sections = ["EXECUTION METADATA","SCAN SCOPE","SCAN STATISTICS","DECISION RESULTS",
                "CONTROL TEST RESULT","DECISION FILTER CHAIN","EXCLUDED SOURCES"]
    for row in ws.iter_rows(max_col=1):
        for cell in row:
            v = str(cell.value or "")
            if any(s in v for s in sections):
                cell.font = Font(bold=True, size=12, color="2F5496")
            if "CONTROL TEST RESULT" in v:
                rc = ws.cell(cell.row, 2)
                if "PASS" in str(rc.value or ""): rc.fill = BG_PASS; rc.font = Font(bold=True, color="548235", size=12)
                elif "CONDITIONAL" in str(rc.value or ""): rc.fill = BG_COND; rc.font = Font(bold=True, color="BF8F00", size=12)
                elif "FAIL" in str(rc.value or ""): rc.fill = BG_FAIL; rc.font = Font(bold=True, color="C00000", size=12)
    ws.column_dimensions["A"].width = 45; ws.column_dimensions["B"].width = 70

    # -- Data Sheets --
    _write_sheet(wb.create_sheet("Non-Compliant"), nc, F_RED, "No non-compliant rules — PASS")
    _write_sheet(wb.create_sheet("Requires Review"), rv, F_AMB, "No rules require manual review")
    _write_sheet(wb.create_sheet("Excluded - Compliant"), ex, F_GRN, "No excluded rules")

    # -- Log --
    wl = wb.create_sheet("Execution Log")
    wl.append(["Timestamp","Level","Message"]); _style_hdr(wl, 1, F_GRY)
    for r in logs: wl.append([r["timestamp"], r["level"], r["message"]])
    _auto_w(wl)

    wb.save(out_path)
    with open(out_path, "rb") as f: sha = hashlib.sha256(f.read()).hexdigest()
    logging.info(f"Report: {out_path} | SHA-256: {sha}")
    return sha


# =============================================================================
# EDR CSV EXPORT
# =============================================================================

def export_edr_csv(nc, path, columns=None):
    """Export non-compliant rules as flat CSV for Splunk ingest (illumio_ruleset.csv)."""
    if columns is None:
        columns = ["Scopes","Sources","Destinations","Destination Services","Ruleset"]
    key_map = {"Scopes":"scopes","Sources":"sources","Destinations":"destinations",
               "Destination Services":"services","Ruleset":"ruleset"}
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        for r in nc:
            w.writerow({c: r.get(key_map.get(c, c.lower()), "") for c in columns})
    logging.info(f"EDR CSV: {path} ({len(nc)} rules)")


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description=f"{CONTROL_ID} — Admin VDI Control Test")
    parser.add_argument("--config", "-c", default=DEFAULT_CONFIG_PATH)
    parser.add_argument("--csv", default=None, help="CSV fallback mode")
    parser.add_argument("--output", "-o", default=None, help="Excel report path")
    parser.add_argument("--edr-export", default=None, help="EDR CSV export path")
    parser.add_argument("--version", "-v", action="version", version=f"%(prog)s {__version__}")
    args = parser.parse_args()

    log_cap = setup_logging()
    logging.info("=" * 72)
    logging.info(f"{CONTROL_ID} — Admin VDI Control Test v{__version__}")
    logging.info("=" * 72)

    try:
        config = load_config(args.config)
        pce = config["_pce"]
        comp = config.get("compliance", {})
        logging.info(f"PCE: {pce['fqdn']} | Ports: {fmt_ports(comp.get('restricted_ports',[]))}")

        client = IllumioClient(pce["fqdn"], pce["port"], pce["org_id"], pce["api_user"], pce["api_key"])

        if args.csv:
            mode = f"csv:{args.csv}"
            nc, rv, ex, stats = process_csv(args.csv, config, client)
        else:
            mode = "api"
            nc, rv, ex, stats = process_api(client, config)

        logging.info("=" * 72)
        logging.info(f"  Rules Scanned:       {stats.get('total_rules',0)}")
        logging.info(f"  Extra-Scope:         {stats.get('extra_scope',0)}")
        logging.info(f"  Restricted Ports:    {stats.get('restricted_port_rules',0)}")
        logging.info(f"  Non-Compliant:       {stats.get('non_compliant',0)}")
        logging.info(f"  Requires Review:     {stats.get('needs_review',0)}")
        logging.info(f"  Excluded:            {stats.get('excluded',0)}")
        logging.info("=" * 72)

        nc_ct, rv_ct = stats.get("non_compliant",0), stats.get("needs_review",0)
        if nc_ct == 0 and rv_ct == 0: logging.info("RESULT: PASS")
        elif nc_ct == 0: logging.warning(f"RESULT: CONDITIONAL — {rv_ct} need review")
        else: logging.error(f"RESULT: FAIL — {nc_ct} non-compliant, {rv_ct} need review")

        out = args.output or f"VDI_Control_Test_{CONTROL_ID}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        generate_report(nc, rv, ex, stats, log_cap.records, config, out, mode)

        edr = args.edr_export or config.get("output",{}).get("edr_export_path")
        if edr and nc:
            export_edr_csv(nc, edr, config.get("output",{}).get("edr_columns"))
        elif edr: logging.info("EDR export skipped — no non-compliant rules")

        return 1 if nc_ct > 0 else (3 if rv_ct > 0 else 0)
    except Exception as e:
        logging.critical(f"FATAL: {e}", exc_info=True)
        return 2

if __name__ == "__main__":
    sys.exit(main())
