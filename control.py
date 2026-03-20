#!/usr/bin/env python3
"""
MON.C9.9 — Admin VDI Control Test (Automated)
Illumio Admin VDI Restrictions Compliance Scanner

Based on the proven working export script. Fetches rulesets, labels, label
groups, IP lists, and services from the Illumio API, applies the MON.C9.9
decision filter chain, and outputs a structured Excel audit report.

Control: MON.C9.9 - Admin VDI Restrictions
Guide:   Admin VDI Control Test Execution v1.2

Decision Filter Chain:
  1. Strip non-end-user policy objects from sources
     (permitted IPLs, A-*, E-*, All Workloads — preserve EUC)
  2. No sources remain → Excluded (all sources were permitted)
  3. EUC present → Requires Review (end-user laptops)
  4. Non-excluded IPL- present → Non-Compliant (finding)
  5. Remaining A-*/E-* only → Excluded (app-to-app permitted)
  5b. Remaining R-* only → Excluded (intra-scope server-to-server)
  6. Only label groups remain → Requires Review (edge case)
  7. Everything else → Non-Compliant

Version: 3.0.0
Author: Cybersecurity Tech Ops — Illuminati Team
"""

__version__ = "3.0.0"

import requests
import csv
import logging
import os
import sys
import json
import time
import hashlib
import getpass
import platform
import argparse
from datetime import datetime, timezone
from pathlib import Path

import yaml
import urllib3
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

CONTROL_ID = "MON.C9.9"
CONTROL_NAME = "Admin VDI Restrictions"
DEFAULT_CONFIG = "config.yaml"


# =============================================================================
# LOGGING
# =============================================================================

class LogCapture(logging.Handler):
    """Captures log records for embedding in Excel report."""
    def __init__(self):
        super().__init__()
        self.records = []
    def emit(self, record):
        self.records.append({
            "timestamp": datetime.fromtimestamp(record.created).strftime("%Y-%m-%d %H:%M:%S"),
            "level": record.levelname,
            "message": self.format(record),
        })


def setup_logging(log_file=None):
    if log_file is None:
        log_file = f"illumio_admin_vdi_log_{time.strftime('%d-%m-%Y')}.txt"
    cap = LogCapture()
    cap.setLevel(logging.DEBUG)
    cap.setFormatter(logging.Formatter("%(message)s"))
    fh = logging.FileHandler(log_file, mode="a")
    fh.setLevel(logging.INFO)
    fh.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    root = logging.getLogger()
    root.setLevel(logging.DEBUG)
    root.addHandler(fh)
    root.addHandler(ch)
    root.addHandler(cap)
    return cap


# =============================================================================
# CONFIGURATION
# =============================================================================

def load_config(path=DEFAULT_CONFIG):
    config = {}
    if Path(path).exists():
        with open(path) as f:
            config = yaml.safe_load(f) or {}
        logging.info(f"Loaded configuration from {path}")
    else:
        logging.warning(f"Config file not found: {path}")

    pce = config.get("pce", {})
    config["_pce"] = {
        "fqdn": os.environ.get("ILLUMIO_FQDN", pce.get("fqdn", "")),
        "port": os.environ.get("ILLUMIO_PORT", str(pce.get("port", "443"))),
        "org_id": os.environ.get("ILLUMIO_ORG_ID", str(pce.get("org_id", ""))),
        "api_user": os.environ.get("ILLUMIO_API_USER", pce.get("api_user", "")),
        "api_key": os.environ.get("ILLUMIO_API_KEY", pce.get("api_key", "")),
    }
    missing = [k for k in ("fqdn", "org_id", "api_user", "api_key") if not config["_pce"][k]]
    if missing:
        raise ValueError(f"Missing PCE config: {', '.join(missing)}")
    return config


# =============================================================================
# ILLUMIO API — PROVEN FETCH LOGIC
# =============================================================================

def fetch_items(url, api_user, api_key):
    """Fetch all items from an Illumio API endpoint (max_results in URL)."""
    logging.info(f"Fetching items from URL: {url}")
    try:
        resp = requests.get(url, verify=False,
                            headers={"Accept": "application/json"},
                            auth=HTTPBasicAuth(api_user, api_key),
                            timeout=120)
        resp.raise_for_status()
        items = resp.json()
        if isinstance(items, dict) and "items" in items:
            items = items["items"]
        logging.info(f"Successfully fetched {len(items)} items from {url}")
        return items
    except requests.exceptions.RequestException as e:
        logging.error(f"Error fetching items from {url}: {e}")
        raise


def build_label_map(base_url, api_user, api_key):
    """Build href → label value map. Only stores A-, R-, E-, L- prefixed labels."""
    logging.info("Building label map...")
    labels = fetch_items(f"{base_url}/labels?max_results=100000", api_user, api_key)
    allowed_prefixes = ("A-", "R-", "E-", "L-")
    label_map = {}
    for label in labels:
        val = label.get("value", "")
        if any(val.startswith(p) for p in allowed_prefixes):
            label_map[label["href"]] = val
    return label_map


def build_label_group_map(base_url, api_user, api_key):
    """Build label_value → set of group names. Uses draft endpoint."""
    logging.info("Building label group map...")
    label_groups = fetch_items(
        f"{base_url}/sec_policy/draft/label_groups?max_results=100000",
        api_user, api_key
    )
    lv_to_groups = {}
    for lg in label_groups:
        group_name = lg.get("name", "")
        for lbl in lg.get("labels", []):
            val = lbl.get("value", "")
            if val:
                lv_to_groups.setdefault(val, set()).add(group_name)
    return lv_to_groups


def build_ip_list_map(base_url, api_user, api_key):
    """Build href → IP list name. Uses draft endpoint, maps both draft and active hrefs."""
    logging.info("Building IP list map...")
    ip_lists = fetch_items(
        f"{base_url}/sec_policy/draft/ip_lists?max_results=100000",
        api_user, api_key
    )
    ip_map = {}
    for ip in ip_lists:
        href = ip["href"]
        name = ip.get("name", "")
        ip_map[href] = name
        if "/draft/ip_lists/" in href:
            ip_map[href.replace("/draft/ip_lists/", "/active/ip_lists/")] = name
    return ip_map


def build_service_map(base_url, api_user, api_key, restricted_ports):
    """Build href → {name, ports, has_restricted}. Active endpoint."""
    logging.info("Building service map...")
    services = fetch_items(
        f"{base_url}/sec_policy/active/services?max_results=100000",
        api_user, api_key
    )
    svc_map = {}
    for svc in services:
        name = svc.get("name", "")
        ports = []
        for sp in svc.get("service_ports", []):
            if "port" in sp and "to_port" in sp:
                ports.append(f"{int(sp['port'])}-{int(sp['to_port'])}")
            elif "port" in sp:
                ports.append(str(int(sp["port"])))
        has_restricted = any(p in restricted_ports for p in ports)
        # Also check port ranges
        if not has_restricted:
            for p in ports:
                if "-" in p:
                    lo, hi = p.split("-")
                    if any(int(lo) <= int(rp) <= int(hi) for rp in restricted_ports):
                        has_restricted = True
                        break
        svc_map[svc["href"]] = {"name": name, "ports": ports, "has_restricted": has_restricted}
    rcount = sum(1 for v in svc_map.values() if v["has_restricted"])
    logging.info(f"Scanning services for restricted ports: {restricted_ports}")
    logging.info(f"Found {rcount} services with restricted ports")
    return svc_map


def fetch_rulesets(base_url, api_user, api_key):
    logging.info("Fetching all rulesets...")
    return fetch_items(f"{base_url}/sec_policy/active/rule_sets?max_results=100000", api_user, api_key)


# =============================================================================
# RESOLVE ACTORS — MATCHING PROVEN LOGIC
# =============================================================================

def resolve_sources(consumers, label_map, ip_list_map, lv_to_groups):
    """Resolve rule consumers to source names and label groups. Matches working script."""
    sources = []
    source_lgs = set()
    for src in consumers:
        if "actors" in src:
            if src["actors"] == "ams":
                sources.append("All Workloads")
                logging.info("Rule source includes All Workloads (ams).")
            else:
                sources.append(f"actors:{src['actors']}")
        if "ip_list" in src:
            href = src["ip_list"].get("href", "")
            name = ip_list_map.get(href, "")
            sources.append(name if name else href)
        if "label" in src:
            href = src["label"].get("href", "") if src["label"] else ""
            val = label_map.get(href, "") if href else ""
            if val:
                sources.append(val)
                for g in lv_to_groups.get(val, []):
                    source_lgs.add(g)
        if "label_group" in src:
            # Label groups resolved via label membership, not directly
            pass
    return sources, source_lgs


def resolve_destinations(providers, label_map, ip_list_map, lv_to_groups):
    """Resolve rule providers to destination names and label groups."""
    dests = []
    dest_lgs = set()
    for dst in providers:
        if "actors" in dst:
            if dst["actors"] == "ams":
                dests.append("All Workloads")
            else:
                dests.append(f"actors:{dst['actors']}")
        if "ip_list" in dst:
            href = dst["ip_list"].get("href", "")
            name = ip_list_map.get(href, "")
            dests.append(name if name else href)
        if "label" in dst:
            href = dst["label"].get("href", "") if dst["label"] else ""
            val = label_map.get(href, "") if href else ""
            if val:
                dests.append(val)
                for g in lv_to_groups.get(val, []):
                    dest_lgs.add(g)
        if "label_group" in dst:
            pass
    return dests, dest_lgs


def resolve_services(ingress_services, svc_map):
    """Resolve services, returning only those with restricted ports."""
    svc_names = []
    for svc in ingress_services:
        info = svc_map.get(svc.get("href", ""), None)
        if info and info["has_restricted"]:
            svc_names.append(info["name"])
    return svc_names


# =============================================================================
# DECISION FILTER ENGINE
# =============================================================================

def matches_pattern(value, patterns):
    """Check if value matches any pattern (supports trailing * wildcard)."""
    for p in patterns:
        if p.endswith("*") and value.startswith(p[:-1]):
            return True
        elif value == p:
            return True
    return False


def apply_decision_filters(sources_list, config):
    """
    Apply the full MON.C9.9 decision filter chain to a rule's source list.

    Args:
        sources_list: list of source name strings (already resolved)
        config: full config dict

    Returns:
        (decision_str, reason_str, remaining_sources_list)
    """
    comp = config.get("compliance", {})
    excluded_ipls = comp.get("excluded_sources", [])
    excluded_labels = comp.get("excluded_labels", [])
    excluded_lgs = comp.get("excluded_label_groups", [])
    euc_patterns = comp.get("euc_patterns", ["A-END_USER_COMPUTE_*"])

    remaining = []
    removed = []

    for src in sources_list:
        # Preserve EUC BEFORE A-* exclusion
        if any(matches_pattern(src, [p]) for p in euc_patterns):
            remaining.append(src)
            continue

        # Excluded labels (A-*, E-*, R-METTLE-CI, etc.)
        if matches_pattern(src, excluded_labels):
            removed.append(src)
            continue

        # Excluded IPLs
        if matches_pattern(src, excluded_ipls):
            removed.append(src)
            continue

        # Excluded label groups
        if matches_pattern(src, excluded_lgs):
            removed.append(src)
            continue

        # All Workloads
        if src == "All Workloads":
            removed.append(src)
            continue

        remaining.append(src)

    # Step 2: No remaining → Excluded
    if not remaining:
        return ("Exclude – No Remaining Sources With Access",
                f"All sources permitted: {'; '.join(removed)}",
                [])

    # Step 3: EUC → Review
    euc = [s for s in remaining if any(matches_pattern(s, [p]) for p in euc_patterns)]
    if euc:
        return ("Keep for Review – Contains A-END_USER_COMPUTE_(EUC)",
                f"EUC: {'; '.join(euc)} — end-user laptops, review for access outside IPL-ADMIN_VDI",
                remaining)

    # Step 4: Non-excluded IPL- → NON-COMPLIANT (any non-excluded IPL is a finding)
    ipls = [s for s in remaining if s.startswith("IPL-")]
    if ipls:
        return ("Non-Compliant – Non-Excluded IP List on Restricted Port",
                f"Non-excluded IPL(s): {'; '.join(ipls)}",
                remaining)

    # Step 5: All A-*/E-* → Excluded (app-to-app / env safety net)
    if all(s.startswith("A-") or s.startswith("E-") for s in remaining):
        return ("Exclude – Application to Application Traffic is Permitted",
                f"App/Env traffic: {'; '.join(remaining)}",
                [])

    # Step 5b: All R-* role labels → Excluded (intra-scope server-to-server)
    # If every remaining source is a role label, this is intra-scope traffic
    # between servers within the same application scope (e.g., R-APPLICATION
    # → R-DATABASE). This is not external end-user access.
    if all(s.startswith("R-") for s in remaining):
        return ("Exclude – Intra-Scope Role-to-Role Traffic",
                f"All sources are role labels (intra-scope): {'; '.join(remaining)}",
                [])

    # Step 6: Only label groups → Review
    if all(s.startswith("LG-") for s in remaining):
        return ("Keep for Review – Edge Case (Label Group Only)",
                f"Only label groups: {'; '.join(remaining)}",
                remaining)

    # Step 7: Non-compliant
    return ("Non-Compliant – Unpermitted Source on Restricted Port",
            f"Non-permitted: {'; '.join(remaining)}",
            remaining)


# =============================================================================
# MAIN PROCESSING
# =============================================================================

def process_rulesets(config, target_cis=None):
    """
    Main processing loop. Fetches all data from PCE, iterates rulesets/rules,
    applies filters, and returns categorized results.
    """
    pce = config["_pce"]
    comp = config.get("compliance", {})
    restricted_ports = set(str(p) for p in comp.get("restricted_ports", ["22", "23", "3389", "7389"]))

    base_url = f"https://{pce['fqdn']}:{pce['port']}/api/v2/orgs/{pce['org_id']}"
    api_user, api_key = pce["api_user"], pce["api_key"]

    # Build all lookup maps (proven logic)
    label_map = build_label_map(base_url, api_user, api_key)
    lv_to_groups = build_label_group_map(base_url, api_user, api_key)
    ip_list_map = build_ip_list_map(base_url, api_user, api_key)
    svc_map = build_service_map(base_url, api_user, api_key, restricted_ports)
    rulesets = fetch_rulesets(base_url, api_user, api_key)

    logging.info(f"Built href lookup: {len(label_map) + len(ip_list_map)} entries")

    non_compliant, needs_review, excluded, blank_rules = [], [], [], []
    stats = {
        "total_rulesets": len(rulesets), "production_rulesets": 0,
        "total_rules": 0, "extra_scope": 0, "restricted_port_rules": 0,
        "non_compliant": 0, "needs_review": 0, "excluded": 0,
        "skip_non_epd": 0, "skip_intra": 0, "skip_disabled": 0,
        "skip_no_restricted": 0, "blank_rules": 0,
    }

    for ruleset in rulesets:
        policy_name = ruleset.get("name", "")
        # Resolve scopes to label values (no prefixes)
        scopes_flat = []
        for scope_group in ruleset.get("scopes", []):
            for scope in scope_group:
                label = scope.get("label", {})
                href = label.get("href", "") if label else ""
                val = label_map.get(href, "") if href else ""
                if val:
                    scopes_flat.append(val)
        scopes_str = "; ".join(scopes_flat)

        rules = ruleset.get("rules", [])
        logging.info(f"Processing ruleset: {policy_name} with {len(rules)} rules.")

        for rule in rules:
            stats["total_rules"] += 1

            # Resolve sources, destinations, services
            sources_list, source_lgs = resolve_sources(
                rule.get("consumers", []), label_map, ip_list_map, lv_to_groups
            )
            dests_list, dest_lgs = resolve_destinations(
                rule.get("providers", []), label_map, ip_list_map, lv_to_groups
            )
            svc_names = resolve_services(rule.get("ingress_services", []), svc_map)

            sources_str = "; ".join(sources_list)
            dests_str = "; ".join(dests_list)
            svc_str = "; ".join(svc_names)
            source_lgs_str = "; ".join(sorted(source_lgs))
            dest_lgs_str = "; ".join(sorted(dest_lgs))

            # Track blank sources/destinations
            if not sources_str or not dests_str:
                blank_rules.append({
                    "policy_name": policy_name, "scopes": scopes_str,
                    "rule_href": rule.get("href", ""),
                })
                stats["blank_rules"] += 1
                logging.warning(f"Rule with blank sources or destinations in ruleset: {policy_name} | {scopes_str}")

            # --- FILTERS (matching working script order) ---
            enabled = rule.get("enabled")
            if enabled is None:
                enabled = ruleset.get("enabled")
            if not enabled:
                logging.info("Rule skipped: not enabled.")
                stats["skip_disabled"] += 1
                continue

            if not svc_names:
                logging.info("Rule skipped: no restricted services.")
                stats["skip_no_restricted"] += 1
                continue

            if "E-PD" not in scopes_str:
                logging.info("Rule skipped: not in E-PD scope.")
                stats["skip_non_epd"] += 1
                continue

            stats["production_rulesets"] += 1  # counted per-rule that passes E-PD filter
            stats["extra_scope"] += 1
            stats["restricted_port_rules"] += 1

            # Build the rule entry
            entry = {
                "ruleset": policy_name,
                "scopes": scopes_str,
                "rule_href": rule.get("href", ""),
                "sources": sources_str,
                "sources_remaining": "",
                "source_label_groups": source_lgs_str,
                "destinations": dests_str,
                "dest_label_groups": dest_lgs_str,
                "services": svc_str,
                "decision": "",
                "reason": "",
            }

            # Target CI filtering (if specified)
            if target_cis:
                if not any(ci in scopes_str for ci in target_cis):
                    # Not in target scope — still process but don't flag as target
                    entry["target_ci"] = ""
                else:
                    entry["target_ci"] = next(ci for ci in target_cis if ci in scopes_str)
            else:
                entry["target_ci"] = ""

            # Apply decision filter chain
            decision, reason, remaining = apply_decision_filters(sources_list, config)
            entry["decision"] = decision
            entry["reason"] = reason
            entry["sources_remaining"] = "; ".join(remaining) if remaining else "(none)"

            if decision.startswith("Exclude"):
                logging.info(f"Rule skipped: {decision}")
                excluded.append(entry)
                stats["excluded"] += 1
            elif decision.startswith("Keep"):
                logging.info(f"Rule flagged: {decision}")
                needs_review.append(entry)
                stats["needs_review"] += 1
            else:
                logging.warning(f"NON-COMPLIANT in {policy_name}: {reason}")
                non_compliant.append(entry)
                stats["non_compliant"] += 1

    # Write blank rules JSON
    if blank_rules:
        blank_path = os.path.join(os.path.dirname(__file__) or ".", "illumio_blank_rules.json")
        with open(blank_path, "w") as jf:
            json.dump(blank_rules, jf, indent=2)
        logging.info(f"Wrote {len(blank_rules)} blank rules to {blank_path}")

    return non_compliant, needs_review, excluded, stats


# =============================================================================
# EXCEL REPORT
# =============================================================================

HF = Font(bold=True, color="FFFFFF", size=11)
F_RED = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
F_AMB = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
F_GRN = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
F_GRY = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
BG_PASS = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BG_FAIL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
BG_COND = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

RULE_HDRS = [
    "Ruleset", "Scopes", "Rule HREF",
    "Sources (Original)", "Sources (After Filter)", "Source Label Groups",
    "Destinations", "Dest Label Groups", "Destination Services",
    "Decision Filter", "Decision Reason",
]


def _style_hdr(ws, row, fill):
    for c in ws[row]:
        c.font = HF; c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN


def _auto_w(ws, mx=55):
    for i, col in enumerate(ws.columns, 1):
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(w + 2, mx)


def _write_rules(ws, rules, fill, empty_msg):
    ws.append(RULE_HDRS)
    _style_hdr(ws, 1, fill)
    if not rules:
        ws.append([empty_msg] + [""] * (len(RULE_HDRS) - 1))
        ws.cell(2, 1).font = Font(italic=True, color="808080")
    else:
        for r in rules:
            ws.append([
                r.get("ruleset", ""), r.get("scopes", ""), r.get("rule_href", ""),
                r.get("sources", ""), r.get("sources_remaining", ""),
                r.get("source_label_groups", ""),
                r.get("destinations", ""), r.get("dest_label_groups", ""),
                r.get("services", ""),
                r.get("decision", ""), r.get("reason", ""),
            ])
    _auto_w(ws)


def generate_report(nc, rv, ex, stats, logs, config, out_path):
    wb = Workbook()
    pce = config["_pce"]
    comp = config.get("compliance", {})
    now_utc = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    nc_ct, rv_ct = stats.get("non_compliant", 0), stats.get("needs_review", 0)

    if nc_ct == 0 and rv_ct == 0:
        result = "PASS — No non-compliant or review-pending rules"
    elif nc_ct == 0:
        result = f"CONDITIONAL — {rv_ct} rule(s) require manual review"
    else:
        result = f"FAIL — {nc_ct} non-compliant, {rv_ct} require review"

    # --- Audit Summary ---
    ws = wb.active
    ws.title = "Audit Summary"
    rows = [
        (f"{CONTROL_ID} — {CONTROL_NAME} — Control Test Report", ""),
        ("", ""),
        ("EXECUTION METADATA", ""),
        ("Script Version", __version__),
        ("Control", f"{CONTROL_ID} — {CONTROL_NAME}"),
        ("Timestamp", now_utc),
        ("Executed By", getpass.getuser()),
        ("Hostname", platform.node()),
        ("PCE Target", pce["fqdn"]),
        ("Org ID", pce["org_id"]),
        ("Data Source", "Illumio API (Direct)"),
        ("", ""),
        ("SCAN SCOPE", ""),
        ("Environment", "E-PD (Production)"),
        ("Rule Scope", "Extra-Scope (rules with restricted services in E-PD)"),
        ("Rule Status", "Enabled Only"),
        ("Restricted Ports", ", ".join(str(p) for p in comp.get("restricted_ports", []))),
        ("", ""),
        ("SCAN STATISTICS", ""),
        ("Total Rulesets Fetched", stats.get("total_rulesets", 0)),
        ("Rules in E-PD with Restricted Services", stats.get("restricted_port_rules", 0)),
        ("Total Rules Scanned", stats.get("total_rules", 0)),
        ("Skipped (Not E-PD)", stats.get("skip_non_epd", 0)),
        ("Skipped (Not Enabled)", stats.get("skip_disabled", 0)),
        ("Skipped (No Restricted Services)", stats.get("skip_no_restricted", 0)),
        ("Blank Sources/Destinations", stats.get("blank_rules", 0)),
        ("", ""),
        ("DECISION RESULTS", ""),
        ("Non-Compliant", nc_ct),
        ("Requires Manual Review", rv_ct),
        ("Excluded / Compliant", stats.get("excluded", 0)),
        ("", ""),
        ("CONTROL TEST RESULT", result),
        ("", ""),
        ("DECISION FILTER CHAIN", ""),
        ("Step 1", "Strip permitted: excluded IPLs, A-*, E-*, All Workloads, specific labels/LGs (preserve EUC)"),
        ("Step 2", "No sources remain → Excluded (all sources were permitted)"),
        ("Step 3", "A-END_USER_COMPUTE_(EUC) present → Requires Review"),
        ("Step 4", "Non-excluded IPL- sources remain → Non-Compliant (finding)"),
        ("Step 5", "All remaining A-*/E-* → Excluded (app-to-app permitted)"),
        ("Step 5b", "All remaining R-* role labels → Excluded (intra-scope server-to-server)"),
        ("Step 6", "Only label groups remain → Requires Review (edge case)"),
        ("Step 7", "Everything else (R-*, other) → Non-Compliant"),
        ("", ""),
        ("EXCLUDED SOURCES (config.yaml)", ""),
    ]
    for s in comp.get("excluded_sources", []):
        rows.append((f"  IPL: {s}", ""))
    for s in comp.get("excluded_labels", []):
        rows.append((f"  Label: {s}", ""))
    for s in comp.get("excluded_label_groups", []):
        rows.append((f"  LG: {s}", ""))
    for s in comp.get("euc_patterns", []):
        rows.append((f"  EUC (Review): {s}", ""))

    for r in rows:
        ws.append(r)

    # Style
    ws["A1"].font = Font(bold=True, size=14)
    sections = ["EXECUTION METADATA", "SCAN SCOPE", "SCAN STATISTICS",
                "DECISION RESULTS", "CONTROL TEST RESULT",
                "DECISION FILTER CHAIN", "EXCLUDED SOURCES"]
    for row in ws.iter_rows(max_col=1):
        for cell in row:
            v = str(cell.value or "")
            if any(s in v for s in sections):
                cell.font = Font(bold=True, size=12, color="2F5496")
            if "CONTROL TEST RESULT" in v:
                rc = ws.cell(cell.row, 2)
                rv_str = str(rc.value or "")
                if "PASS" in rv_str:
                    rc.fill = BG_PASS; rc.font = Font(bold=True, color="548235", size=12)
                elif "CONDITIONAL" in rv_str:
                    rc.fill = BG_COND; rc.font = Font(bold=True, color="BF8F00", size=12)
                elif "FAIL" in rv_str:
                    rc.fill = BG_FAIL; rc.font = Font(bold=True, color="C00000", size=12)
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 70

    # --- Data Sheets ---
    _write_rules(wb.create_sheet("Non-Compliant"), nc, F_RED, "No non-compliant rules — PASS")
    _write_rules(wb.create_sheet("Requires Review"), rv, F_AMB, "No rules require manual review")
    _write_rules(wb.create_sheet("Excluded - Compliant"), ex, F_GRN, "No excluded rules")

    # --- Execution Log ---
    wl = wb.create_sheet("Execution Log")
    wl.append(["Timestamp", "Level", "Message"])
    _style_hdr(wl, 1, F_GRY)
    for rec in logs:
        wl.append([rec["timestamp"], rec["level"], rec["message"]])
    _auto_w(wl)

    wb.save(out_path)
    with open(out_path, "rb") as f:
        sha = hashlib.sha256(f.read()).hexdigest()
    logging.info(f"Report: {out_path} | SHA-256: {sha}")
    return sha


# =============================================================================
# EDR CSV EXPORT
# =============================================================================

def export_edr_csv(nc, path):
    """Export non-compliant rules as flat CSV for Splunk ingest."""
    cols = ["Scopes", "Sources", "Destinations", "Destination Services", "Ruleset"]
    kmap = {"Scopes": "scopes", "Sources": "sources", "Destinations": "destinations",
            "Destination Services": "services", "Ruleset": "ruleset"}
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in nc:
            w.writerow({c: r.get(kmap[c], "") for c in cols})
    logging.info(f"EDR CSV: {path} ({len(nc)} rules)")


# =============================================================================
# WORKLOADER VALIDATION (optional backup via brian1917/workloader)
# =============================================================================

def run_workloader_validation(workloader_path, config, nc_rules, script_dir=None):
    """
    Run workloader rule-export as a validation/backup source.

    1. Calls workloader rule-export --policy-version active --expand-svcs
    2. Parses the CSV output
    3. Applies the same E-PD + restricted port filters
    4. Compares against our script's non-compliant results
    5. Logs discrepancies

    Args:
        workloader_path: Path to workloader binary
        config: Full config dict
        nc_rules: List of non-compliant rule dicts from our script
        script_dir: Directory to save workloader output (default: cwd)
    """
    import subprocess

    if script_dir is None:
        script_dir = os.path.dirname(os.path.abspath(__file__)) or "."

    comp = config.get("compliance", {})
    restricted_ports = set(str(p) for p in comp.get("restricted_ports", ["22", "23", "3389", "7389"]))

    # Verify workloader binary exists
    wl_bin = os.path.abspath(workloader_path)
    if not os.path.isfile(wl_bin):
        # Try as command in PATH
        import shutil
        wl_bin = shutil.which(workloader_path)
        if not wl_bin:
            logging.error(f"Workloader binary not found: {workloader_path}")
            return None

    logging.info("=" * 72)
    logging.info("WORKLOADER VALIDATION")
    logging.info(f"Binary: {wl_bin}")
    logging.info("=" * 72)

    # Build output filename
    wl_output = os.path.join(script_dir, f"workloader_rule_export_{time.strftime('%Y%m%d_%H%M%S')}.csv")

    # Build workloader command
    # Flags from: workloader rule-export --help
    #   --policy-version active  (export active/provisioned rules)
    #   --expand-svcs            (show ports/protocols instead of service hrefs)
    #   --no-href                (clean output, no href column)
    #   --output-file <path>     (specify output CSV path)
    #   --out csv                (output format)
    cmd = [
        wl_bin, "rule-export",
        "--policy-version", "active",
        "--expand-svcs",
        "--no-href",
        "--output-file", wl_output,
        "--out", "csv",
    ]

    # Add config file if specified in our config (path to pce.yaml)
    wl_config = config.get("workloader", {}).get("config_file")
    if wl_config:
        cmd.extend(["--config-file", wl_config])

    # Add PCE override if multiple PCEs configured in pce.yaml
    wl_pce = config.get("workloader", {}).get("pce_name")
    if wl_pce:
        cmd.extend(["--pce", wl_pce])

    logging.info(f"Running: {' '.join(cmd)}")

    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=300
        )

        if result.returncode != 0:
            logging.error(f"Workloader exited with code {result.returncode}")
            if result.stderr:
                logging.error(f"Workloader stderr: {result.stderr[:500]}")
            if result.stdout:
                logging.info(f"Workloader stdout: {result.stdout[:500]}")
            return None

        logging.info(f"Workloader export complete: {wl_output}")

    except FileNotFoundError:
        logging.error(f"Could not execute workloader: {wl_bin}")
        return None
    except subprocess.TimeoutExpired:
        logging.error("Workloader timed out after 300 seconds")
        return None
    except Exception as e:
        logging.error(f"Workloader error: {e}")
        return None

    # --- Parse workloader output and apply our filters ---
    if not os.path.isfile(wl_output):
        logging.error(f"Workloader output file not found: {wl_output}")
        return None

    try:
        wl_rules = []
        with open(wl_output, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            wl_headers = reader.fieldnames
            logging.info(f"Workloader CSV columns: {wl_headers}")
            for row in reader:
                wl_rules.append(row)

        logging.info(f"Workloader exported {len(wl_rules)} total rules")

        # Apply our filters to workloader output
        # Workloader column names vary but typically include:
        #   ruleset_name, ruleset_scope, src_labels, src_iplists,
        #   dst_labels, dst_iplists, services (or expanded port columns)
        # With --expand-svcs the services show as ports/protocols

        wl_filtered = []
        for row in wl_rules:
            # Get scope — try common column names
            scopes = (row.get("ruleset_scope", "")
                      or row.get("Scopes", "")
                      or row.get("scopes", ""))

            # E-PD filter
            if "E-PD" not in scopes:
                continue

            # Get services — try common column names
            services = (row.get("services", "")
                        or row.get("Destination Services", "")
                        or row.get("ingress_services", ""))

            # Check for restricted ports in services
            has_restricted = False
            for rp in restricted_ports:
                if rp in services:
                    has_restricted = True
                    break

            if not has_restricted:
                continue

            # Get sources
            sources = (row.get("src_labels", "") + ";" +
                       row.get("src_iplists", "") +
                       row.get("Sources", ""))

            wl_filtered.append({
                "ruleset": row.get("ruleset_name", row.get("Ruleset", "")),
                "scopes": scopes,
                "sources": sources,
                "services": services,
            })

        logging.info(f"Workloader rules after E-PD + restricted port filter: {len(wl_filtered)}")

        # --- Compare against our results ---
        our_rulesets = set(r["ruleset"] for r in nc_rules)
        wl_rulesets = set(r["ruleset"] for r in wl_filtered)

        only_in_ours = our_rulesets - wl_rulesets
        only_in_wl = wl_rulesets - our_rulesets
        in_both = our_rulesets & wl_rulesets

        logging.info(f"COMPARISON:")
        logging.info(f"  Our non-compliant rulesets:        {len(our_rulesets)}")
        logging.info(f"  Workloader filtered rulesets:      {len(wl_rulesets)}")
        logging.info(f"  In both:                           {len(in_both)}")

        if only_in_ours:
            logging.warning(f"  In our results but NOT workloader: {len(only_in_ours)}")
            for rs in sorted(only_in_ours):
                logging.warning(f"    - {rs}")

        if only_in_wl:
            logging.warning(f"  In workloader but NOT our results: {len(only_in_wl)}")
            for rs in sorted(only_in_wl):
                logging.warning(f"    - {rs}")

        if not only_in_ours and not only_in_wl:
            logging.info("  VALIDATION PASSED — results match")

        return {
            "wl_total": len(wl_rules),
            "wl_filtered": len(wl_filtered),
            "in_both": len(in_both),
            "only_ours": sorted(only_in_ours),
            "only_wl": sorted(only_in_wl),
            "wl_output_file": wl_output,
        }

    except Exception as e:
        logging.error(f"Error parsing workloader output: {e}")
        return None


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description=f"{CONTROL_ID} — Admin VDI Control Test v{__version__}"
    )
    parser.add_argument("--config", "-c", default=DEFAULT_CONFIG)
    parser.add_argument("--output", "-o", default=None, help="Excel report path")
    parser.add_argument("--edr-export", default=None, help="EDR CSV path for Splunk")
    parser.add_argument("--target-cis", default=None,
                        help="Comma-separated Target CIs to filter by")
    parser.add_argument("--workloader", default=None,
                        help="Path to workloader binary for validation "
                             "(e.g., ./workloader or C:\\Scripts\\workloader.exe)")
    parser.add_argument("--version", "-v", action="version",
                        version=f"%(prog)s {__version__}")
    args = parser.parse_args()

    log_cap = setup_logging()
    logging.info("=" * 72)
    logging.info(f"{CONTROL_ID} — Admin VDI Control Test v{__version__}")
    logging.info("=" * 72)

    try:
        config = load_config(args.config)
        pce = config["_pce"]
        comp = config.get("compliance", {})

        logging.info(f"PCE: {pce['fqdn']}:{pce['port']} | Org: {pce['org_id']}")
        logging.info(f"Restricted Ports: {comp.get('restricted_ports', [])}")
        excl = (len(comp.get("excluded_sources", []))
                + len(comp.get("excluded_labels", []))
                + len(comp.get("excluded_label_groups", [])))
        logging.info(f"Excluded source patterns: {excl}")
        logging.info("Mode: Direct API pull from PCE")

        # Parse target CIs
        target_cis = None
        if args.target_cis:
            target_cis = [ci.strip() for ci in args.target_cis.split(",") if ci.strip()]
            logging.info(f"Target CIs: {target_cis}")

        # Process
        nc, rv, ex, stats = process_rulesets(config, target_cis)

        # Results
        logging.info("=" * 72)
        logging.info("RESULTS")
        logging.info(f"  Total Rules Scanned:          {stats.get('total_rules', 0)}")
        logging.info(f"  E-PD + Restricted Services:   {stats.get('restricted_port_rules', 0)}")
        logging.info(f"  Non-Compliant:                {stats.get('non_compliant', 0)}")
        logging.info(f"  Requires Review:              {stats.get('needs_review', 0)}")
        logging.info(f"  Excluded:                     {stats.get('excluded', 0)}")
        logging.info("=" * 72)

        nc_ct = stats.get("non_compliant", 0)
        rv_ct = stats.get("needs_review", 0)
        if nc_ct == 0 and rv_ct == 0:
            logging.info("RESULT: PASS")
        elif nc_ct == 0:
            logging.warning(f"RESULT: CONDITIONAL — {rv_ct} need review")
        else:
            logging.error(f"RESULT: FAIL — {nc_ct} non-compliant, {rv_ct} need review")

        # Generate report
        out = args.output or f"VDI_Control_Test_{CONTROL_ID}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        generate_report(nc, rv, ex, stats, log_cap.records, config, out)

        # EDR export
        edr = args.edr_export or config.get("output", {}).get("edr_export_path")
        if edr and nc:
            export_edr_csv(nc, edr)
        elif edr:
            logging.info("EDR export skipped — no non-compliant rules")

        # Workloader validation (optional)
        if args.workloader:
            wl_result = run_workloader_validation(
                workloader_path=args.workloader,
                config=config,
                nc_rules=nc,
                script_dir=os.path.dirname(os.path.abspath(out)),
            )
            if wl_result:
                logging.info(f"Workloader export saved: {wl_result['wl_output_file']}")
                if wl_result["only_ours"] or wl_result["only_wl"]:
                    logging.warning("Workloader validation found discrepancies — review logs")
                else:
                    logging.info("Workloader validation PASSED")

        logging.info("Script completed successfully.")
        return 1 if nc_ct > 0 else (3 if rv_ct > 0 else 0)

    except Exception as e:
        logging.critical(f"FATAL: {e}", exc_info=True)
        return 2


if __name__ == "__main__":
    sys.exit(main())
